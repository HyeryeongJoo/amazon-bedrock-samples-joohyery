# utils.py
"""
번역 문서 생성을 위한 유틸리티 함수들
"""

import os
import base64
import io
import pandas as pd
from PIL import Image
from datetime import datetime
from typing import Dict, List, Optional
import boto3
from botocore.config import Config

# =============================================================================
# 파일 타입 감지 함수들
# =============================================================================

def is_pdf(file_path: str) -> bool:
    """파일이 PDF인지 확인"""
    return file_path.lower().endswith('.pdf')

def is_image(file_path: str) -> bool:
    """파일이 이미지인지 확인"""
    try:
        with Image.open(file_path) as img:
            img.verify()
        return True
    except Exception:
        return False

def get_file_type(file_path: str) -> str:
    """파일 타입 반환 (pdf, image, unknown)"""
    if is_pdf(file_path):
        return 'pdf'
    elif is_image(file_path):
        return 'image'
    else:
        return 'unknown'

# =============================================================================
# PDF 처리 함수들
# =============================================================================

def encode_pdf(pdf_path: str) -> str:
    """PDF 파일을 base64로 인코딩"""
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF 파일을 찾을 수 없습니다: {pdf_path}")
    
    try:
        with open(pdf_path, 'rb') as pdf_file:
            return base64.b64encode(pdf_file.read()).decode('utf-8')
    except Exception as e:
        raise Exception(f"PDF 인코딩 중 오류 발생: {str(e)}")

def process_pdf_with_claude(pdf_path: str, client=None, model_id="us.anthropic.claude-sonnet-4-20250514-v1:0") -> str:
    """Claude 4의 PDF Chat 기능을 사용하여 PDF 텍스트 추출"""
    if client is None:
        config = Config(read_timeout=300)
        client = boto3.client(service_name="bedrock-runtime", region_name="us-west-2", config=config)
    
    try:
        # PDF를 base64로 인코딩
        pdf_base64 = encode_pdf(pdf_path)
        
        # Claude 4 PDF Chat용 시스템 프롬프트
        system_prompt = """You are a professional text extraction specialist. Your task is to carefully analyze PDF documents and extract ALL text content for translation purposes.

## Instructions:
1. Extract ALL visible text from the PDF, including:
   - Headlines and titles
   - Body text and paragraphs
   - Captions and descriptions
   - Contact information (phone numbers, addresses, emails, websites)
   - Small print and disclaimers
   - Menu items, prices, or product listings
   - Date and time information
   - Terms and conditions
   - Table contents and data

2. Organize the extracted text in a logical order:
   - Follow the document flow (page by page, section by section)
   - Maintain the original structure as much as possible
   - Group related content together
   - Clearly separate different sections

3. Present the text in the original language - do NOT translate anything
4. If text is unclear or partially visible, note it as [UNCLEAR: approximate text]
5. Maintain formatting structure where important for context

Your goal is to ensure no text content is missed so that the subsequent translation will be complete and accurate."""
        
        user_prompt = f"""Please extract all text content from this PDF document. 
The document may contain Korean text that needs to be extracted for translation purposes.
Maintain the original structure and organization of the content."""
        
        # Converse API 호출
        response = client.converse(
            modelId=model_id,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "document": {
                                "format": "pdf",
                                "name": os.path.basename(pdf_path),
                                "source": {
                                    "bytes": base64.b64decode(pdf_base64)
                                }
                            }
                        },
                        {
                            "text": user_prompt
                        }
                    ]
                }
            ],
            system=[
                {
                    "text": system_prompt
                }
            ],
            inferenceConfig={
                "maxTokens": 4000,
                "temperature": 0.1
            }
        )
        
        return response['output']['message']['content'][0]['text']
        
    except Exception as e:
        raise Exception(f"PDF 처리 중 오류 발생: {str(e)}")

# =============================================================================
# 이미지 처리 함수들
# =============================================================================

def get_image_format(image_path: str) -> str:
    """이미지 형식 자동 감지"""
    try:
        with Image.open(image_path) as img:
            format_map = {
                'JPEG': 'jpeg',
                'PNG': 'png',
                'GIF': 'gif',
                'BMP': 'bmp',
                'WEBP': 'webp'
            }
            return format_map.get(img.format, 'jpeg')
    except:
        # 기본값으로 jpeg 반환
        return 'jpeg'

def validate_and_resize_image(image_path: str, max_pixel: int = 8000) -> str:
    """이미지 크기 검증 및 필요시 리사이징 (세로, 가로 중 하나라도 max_pixel 초과하지 않도록)"""
    try:
        # 이미지 열기
        with Image.open(image_path) as img:
            width, height = img.size
            print(f"원본 이미지 크기: {width}x{height}")
            
            # 크기 검증
            if height <= max_pixel and width <= max_pixel:
                print("이미지 크기가 적절합니다.")
                # 원본 이미지를 base64로 인코딩
                with open(image_path, 'rb') as image_file:
                    return base64.b64encode(image_file.read()).decode('utf-8')
            
            # 리사이징 필요
            print(f"이미지 크기가 너무 큽니다. 리사이징 중... (최대: {max_pixel}x{max_pixel})")
            
            # 비율 유지하며 리사이징
            ratio = min(max_pixel / width, max_pixel / height)
            new_width = int(width * ratio)
            new_height = int(height * ratio)
            
            resized_img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
            print(f"리사이징된 이미지 크기: {new_width}x{new_height}")
            
            # 메모리에서 바이트로 변환
            img_buffer = io.BytesIO()
            # 원본 형식 유지 (JPEG, PNG 등)
            format = img.format if img.format else 'JPEG'
            resized_img.save(img_buffer, format=format, quality=95)
            img_buffer.seek(0)
            
            return base64.b64encode(img_buffer.getvalue()).decode('utf-8')
    
    except Exception as e:
        raise Exception(f"이미지 처리 중 오류 발생: {str(e)}")

def encode_image(image_path: str) -> str:
    """이미지를 base64로 인코딩 (크기 검증 포함)"""
    if not os.path.exists(image_path):
        raise FileNotFoundError(f"이미지 파일을 찾을 수 없습니다: {image_path}")
    
    try:
        # 이미지 크기 검증 및 필요시 리사이징
        return validate_and_resize_image(image_path)
    except Exception as e:
        raise Exception(f"이미지 인코딩 중 오류 발생: {str(e)}")

def encode_file(file_path: str) -> tuple:
    """파일 타입에 따라 적절한 인코딩 방법 선택
    
    Returns:
        tuple: (encoded_data, file_type, format)
    """
    file_type = get_file_type(file_path)
    
    if file_type == 'pdf':
        encoded_data = encode_pdf(file_path)
        return encoded_data, 'pdf', 'pdf'
    elif file_type == 'image':
        encoded_data = encode_image(file_path)
        image_format = get_image_format(file_path)
        return encoded_data, 'image', image_format
    else:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {file_path}")

# =============================================================================
# 파일 처리 함수들
# =============================================================================

def read_html_content(html_path: str) -> str:
    """HTML 파일 내용 읽기"""
    with open(html_path, 'r', encoding='utf-8') as file:
        return file.read()

def add_python_path(module_path):
    """파이썬 경로 추가"""
    import sys
    if os.path.abspath(module_path) not in sys.path:
        sys.path.append(os.path.abspath(module_path))
        print(f"python path: {os.path.abspath(module_path)} is added")
    else:
        print(f"python path: {os.path.abspath(module_path)} already exists")
    print("sys.path: ", sys.path)

# =============================================================================
# 번역 문서 생성 함수들
# =============================================================================

def format_translation_document(grouped_texts, 
                               source_language: str = "Korean", 
                               target_language: str = "English") -> pd.DataFrame:
    """
    번역가가 사용할 엑셀 문서 형식으로 데이터를 포맷팅
    """
    
    # 1. 그룹 리스트 추출
    if isinstance(grouped_texts, dict) and 'groups' in grouped_texts:
        groups = grouped_texts['groups']
    elif isinstance(grouped_texts, list):
        groups = grouped_texts
    else:
        groups = [grouped_texts]  # 단일 값인 경우 리스트로 변환
    
    translation_data = []
    
    for i, group in enumerate(groups, 1):
        try:
            group_text = ""
            
            if isinstance(group, dict):
                # 딕셔너리인 경우: {"category": "...", "texts": [...]} 형태
                if 'texts' in group:
                    texts = group['texts']
                    if isinstance(texts, list):
                        group_text = ' | '.join(str(item) for item in texts if str(item).strip())
                    else:
                        group_text = str(texts).strip()
                else:
                    # texts 키가 없는 경우 전체를 문자열로 변환
                    group_text = str(group).strip()
                    
            elif isinstance(group, list):
                # 리스트인 경우: ["텍스트1", "텍스트2", ...] 형태  
                group_text = ' | '.join(str(item) for item in group if str(item).strip())
                
            else:
                # 기타 형태
                group_text = str(group).strip()
            
            # 빈 텍스트는 건너뛰기
            if not group_text:
                continue
                
            translation_data.append({
                'ID': f'T{i:03d}',
                'Category': f'Group_{i}',
                'Priority': 'medium',
                'Location': '',
                'Description': '',
                f'Original_Text_{source_language}': group_text,
                f'Translated_Text_{target_language}': '',
                'Notes': '',
                'Status': 'Pending'
            })
            
        except Exception as e:
            print(f"그룹 {i} 처리 중 오류: {e}")
            continue
    
    if not translation_data:
        print("경고: 처리할 수 있는 텍스트 그룹이 없습니다.")
        # 빈 데이터프레임이라도 구조는 유지
        return pd.DataFrame(columns=[
            'ID', 'Category', 'Priority', 'Location', 'Description', 
            f'Original_Text_{source_language}', f'Translated_Text_{target_language}', 
            'Notes', 'Status'
        ])
    
    return pd.DataFrame(translation_data)


def save_translation_document(df: pd.DataFrame, 
                             filename: str = None, 
                             document_name: str = "document") -> str:
    """
    번역 문서를 final_results 폴더에 엑셀 파일로 저장
    
    Args:
        df: 저장할 데이터프레임
        filename: 파일명 (None인 경우 자동 생성)
        document_name: 원본 문서 이름 (파일명에 포함)
    
    Returns:
        str: 저장된 파일의 전체 경로
    """
    
    # final_results 폴더 생성
    current_dir = os.getcwd()
    final_results_dir = os.path.join(current_dir, 'final_results')
    
    if not os.path.exists(final_results_dir):
        os.makedirs(final_results_dir)
        print(f"'{final_results_dir}' 폴더를 생성했습니다.")
    
    # 파일명 생성
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"translation_document_{document_name}_{timestamp}.xlsx"
    
    # .xlsx 확장자 확인
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # 전체 파일 경로
    file_path = os.path.join(final_results_dir, filename)
    
    # 엑셀 파일로 저장 (스타일 적용)
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # 메인 번역 시트
        df.to_excel(writer, sheet_name='Translation', index=False)
        
        # 워크북과 워크시트 객체 가져오기
        workbook = writer.book
        worksheet = writer.sheets['Translation']
        
        # 스타일 적용
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # 테두리 스타일
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 헤더 행 스타일 적용
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 데이터 행 스타일 적용
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        
        # 열 너비 자동 조정
        column_widths = {
            'A': 8,   # ID
            'B': 15,  # Category
            'C': 10,  # Priority
            'D': 15,  # Location
            'E': 20,  # Description
            'F': 40,  # Original Text
            'G': 40,  # Translated Text
            'H': 20,  # Notes
            'I': 12   # Status
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
        
        # 행 높이 조정
        for row in range(2, worksheet.max_row + 1):
            worksheet.row_dimensions[row].height = 30
    
    print(f"번역 문서가 저장되었습니다: {file_path}")
    return file_path

# =============================================================================
# 통합 처리 함수들
# =============================================================================

def process_document_with_claude(file_path: str, client=None, model_id="us.anthropic.claude-sonnet-4-20250514-v1:0") -> str:
    """문서 타입에 따라 적절한 Claude 모델 기능을 사용하여 텍스트 추출"""
    if client is None:
        config = Config(read_timeout=300)
        client = boto3.client(service_name="bedrock-runtime", region_name="us-west-2", config=config)
    
    file_type = get_file_type(file_path)
    
    if file_type == 'pdf':
        print(f"PDF 파일 감지: Claude 4 PDF Chat 기능 사용")
        return process_pdf_with_claude(file_path, client, model_id)
    elif file_type == 'image':
        print(f"이미지 파일 감지: Claude 4 Vision 기능 사용")
        return process_image_with_claude(file_path, client, model_id)
    else:
        raise ValueError(f"지원하지 않는 파일 형식입니다: {file_path}")

def process_image_with_claude(image_path: str, client=None, model_id="us.anthropic.claude-sonnet-4-20250514-v1:0") -> str:
    """Claude 4의 Vision 기능을 사용하여 이미지 텍스트 추출"""
    if client is None:
        config = Config(read_timeout=300)
        client = boto3.client(service_name="bedrock-runtime", region_name="us-west-2", config=config)
    
    try:
        # 이미지를 base64로 인코딩
        image_base64 = encode_image(image_path)
        image_format = get_image_format(image_path)
        
        # 시스템 프롬프트
        system_prompt = """You are a professional text extraction specialist. Your task is to carefully analyze Korean leaflets/brochures and extract ALL text content for English translation purposes.

## Instructions:
1. Examine the provided Korean leaflet thoroughly
2. Extract ALL visible text, including:
   - Headlines and titles
   - Body text and paragraphs
   - Captions and descriptions
   - Contact information (phone numbers, addresses, emails, websites)
   - Small print and disclaimers
   - Menu items, prices, or product listings
   - Date and time information
   - Terms and conditions

3. DO NOT extract text that appears within images, logos, graphics, or decorative elements
   - Focus only on regular text content that is part of the leaflet's main text layout
   - Skip any text that is embedded in photographs, illustrations, or logo designs

4. Organize the extracted text in a logical order:
   - Start with main headings/titles
   - Follow the visual flow of the leaflet (left to right, top to bottom)
   - Group related content together
   - Clearly separate different sections

5. Present the text in the original Korean language - do NOT translate anything
6. If text is unclear or partially visible, note it as [UNCLEAR: approximate text]
7. Maintain the original formatting structure as much as possible

Your goal is to ensure no regular text content is missed while excluding text within images/graphics, so that the subsequent English translation will be complete and accurate for the main textual content only."""
        
        user_prompt = f"""제공된 리플렛의 텍스트를 추출하세요."""
        
        # Converse API 호출
        response = client.converse(
            modelId=model_id,
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "image": {
                                "format": image_format,
                                "source": {
                                    "bytes": base64.b64decode(image_base64)
                                }
                            }
                        },
                        {
                            "text": user_prompt
                        }
                    ]
                }
            ],
            system=[
                {
                    "text": system_prompt
                }
            ],
            inferenceConfig={
                "maxTokens": 4000,
                "temperature": 0.1
            }
        )
        
        return response['output']['message']['content'][0]['text']
        
    except Exception as e:
        raise Exception(f"이미지 처리 중 오류 발생: {str(e)}")

def create_translation_workflow(grouped_texts, 
                               document_name: str = "document",
                               source_lang: str = "Korean", 
                               target_lang: str = "English") -> str:
    """번역 워크플로우 생성 (이미지와 PDF 모두 지원)"""
    print("번역 문서를 포맷팅 중...")
    df = format_translation_document(grouped_texts, source_lang, target_lang)
    
    print("번역 문서를 저장 중...")
    file_path = save_translation_document(df, document_name=document_name)
    
    print(f"총 {len(df)}개의 텍스트 그룹이 포함되어 있습니다.")
    print(f"\n=== 번역 문서 생성 완료 ===")
    print(f"파일 경로: {file_path}")
    print(f"텍스트 그룹 수: {len(df)}")
    print(f"소스 언어: {source_lang}")
    print(f"타겟 언어: {target_lang}")
    
    return file_path

def check_file_paths(image_path: str, html_path: str) -> tuple:
    """파일 경로 존재 여부 확인"""
    image_exists = os.path.exists(image_path)
    html_exists = os.path.exists(html_path)
    
    print(f"이미지 파일: {image_path} ({'존재' if image_exists else '없음'})")
    print(f"HTML 파일: {html_path} ({'존재' if html_exists else '없음'})")
    
    return image_exists, html_exists

def list_files_in_directory(directory_path: str, extensions: List[str] = None) -> List[str]:
    """디렉토리 내 특정 확장자 파일들 나열"""
    if not os.path.exists(directory_path):
        return []
    
    files = []
    for filename in os.listdir(directory_path):
        if extensions:
            if any(filename.lower().endswith(ext.lower()) for ext in extensions):
                files.append(filename)
        else:
            files.append(filename)
    
    return sorted(files)
